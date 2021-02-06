import cv2
import PIL
import numpy as np
from PIL import Image,ImageDraw,ImageFont
import os
from openpyxl import load_workbook
import imageio
import re
from gtts import gTTS
from playsound import playsound
#from moviepy.editor import *
import html
from google.cloud import texttospeech
from pydub import AudioSegment

from moviepy.video.io.VideoFileClip import VideoFileClip as mpe
 
import ffmpeg
import natsort

import sys

if getattr(sys, 'frozen', False):
    os.environ['SSL_CERT_FILE'] = os.path.join(sys._MEIPASS, 'lib', 'cert.pem')

#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("C:\\ttsInExcel\include\excelFile.xlsx", data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['Sheet1']
max_row = load_ws.max_row

selectedFont =ImageFont.truetype("SCDream7.otf", 35) # 폰트경로과 사이즈 설정
b,g,r,a = 255,255,255,0

count = 1

def combine_audio(vidname, audname, outname, fps=25):
    import moviepy.editor as mpe
    my_clip = mpe.VideoFileClip(vidname)
    audio_background = mpe.AudioFileClip(audname)
    final_clip = my_clip.set_audio(audio_background)
    final_clip.write_videofile(outname,fps=fps)

def ssml_to_audio(ssml_text, outfile, lang):

    # Instantiates a client
    client = texttospeech.TextToSpeechClient()

    # Sets the text input to be synthesized
    synthesis_input = texttospeech.SynthesisInput(ssml=ssml_text)

    # Builds the voice request, selects the language code ("en-US") and
    # the SSML voice gender ("MALE")
    voice = texttospeech.VoiceSelectionParams(
        language_code=lang, ssml_gender=texttospeech.SsmlVoiceGender.FEMALE
    )

    # Selects the type of audio file to return
    audio_config = texttospeech.AudioConfig(
        audio_encoding=texttospeech.AudioEncoding.MP3
    )

    # Performs the text-to-speech request on the text input with the selected
    # voice parameters and audio file type
    response = client.synthesize_speech(
        input=synthesis_input, voice=voice, audio_config=audio_config
    )

    # Writes the synthetic audio to the output file.
    with open(outfile, "wb") as out:
        out.write(response.audio_content)

def text_to_ssml(inputfile, speed):
    
    raw_lines = inputfile

    escaped_lines = html.escape(raw_lines)

    # Convert plaintext to SSML
    # Wait two seconds between each address
    ssml = "<speak><prosody rate='"+ speed + "'>{}</prosody></speak>".format(
        escaped_lines.replace("  ", '<break time="0.5s"/>').replace("# ", '<break time="0.25s"/>').replace("* ", '<break time="2.2s"/>').replace("0 ", '<break time="2s"/>').replace("1 ", '<break time="1.5s"/>').replace("2 ", '<break time="1s"/>').replace("3 ", '<break time="0.7s"/>')
    )

    print(ssml)
 
    # Return the concatenated string of ssml script
    return ssml

print("파일 읽는 중 입니다..")

for i in load_ws.rows:
    try:
        target_image = Image.open('C:\\ttsInExcel\\include\\background.jpg')  #일단 기본배경폼 이미지를 open 합니다.
        draw = ImageDraw.Draw(target_image)
        num = i[1].value
        eng = i[3].value
        kor = i[4].value
        word = eng + " : " + kor

        print()
        print(str(num) + ". " + word)
        
        imgNm = "C:\\ttsInExcel\png" + str(num) + ".png"
        audioNm = "C:\\ttsInExcel\\audio" + str(num) + ".mp3"
        audioKoNm = "C:\\ttsInExcel\\audioKo" + str(num) + ".mp3"
        
        draw.text((80, 50), str(num), font=ImageFont.truetype("SCDream7.otf", 35), fill=(160,161,157,a))

        if len(eng) >= 13 or len(kor) >= 7:
            if len(eng) >= 18 :
                draw.text((130, 310), eng, font=selectedFont, fill=(b,g,r,a))
            elif len(eng) >= 13 :
                draw.text((250, 310), eng, font=selectedFont, fill=(b,g,r,a))        
            elif len(eng) >= 9 : 
                draw.text((330, 310), eng, font=selectedFont, fill=(b,g,r,a))
            else:
                draw.text((400, 310), eng, font=selectedFont, fill=(b,g,r,a))
            
            if len(kor) >= 16 :
                draw.text((720, 310), kor[0:17], font=selectedFont, fill=(b,g,r,a))
                draw.text((720, 410), kor[17:], font=selectedFont, fill=(b,g,r,a))
               
                kor = kor + "3 "
            else :
                draw.text((720, 310), kor, font=selectedFont, fill=(b,g,r,a))
                
                kor = kor + "2 "
            
            if len(eng) <= 4 :
                eng = eng + "* "
            else : 
                eng = eng + "1 " 
            
        else :
            if len(eng) >= 11 :
                draw.text((330, 310), eng, font=selectedFont, fill=(b,g,r,a))
                eng = eng + "2 "
            elif len(eng) <= 9 and len(eng) >= 5:
                draw.text((380, 310), eng, font=selectedFont, fill=(b,g,r,a))
                eng = eng + "1 "
            else :
                draw.text((400, 310), eng, font=selectedFont, fill=(b,g,r,a))
                eng = eng + "0 "
            draw.text((700, 310), kor, font=selectedFont, fill=(b,g,r,a))   

            if len(kor) < 4 :
                kor = kor + "* "
            else : 
                kor = kor + "0 "
        
        target_image.save(imgNm) # 편집된 이미지 저장
        
        ssml = text_to_ssml(eng, "100%")
        if len(kor) >= 10 :
            ssml_ko = text_to_ssml(kor, "110%")
        else:
            ssml_ko = text_to_ssml(kor, "100%")   

        ssml_to_audio(ssml, audioNm, "en-US")
        ssml_to_audio(ssml_ko, audioKoNm, "ko-KR")

        sound1 = AudioSegment.from_mp3(audioNm)
        sound2 = AudioSegment.from_mp3(audioKoNm)

        # sound1, with sound2 appended (use louder instead of sound1 to append the louder version)

        combined = sound1 + sound2

        # save the result
        combined.export("C:\\ttsInExcel\\audio_mixed_sounds" + str(num) + ".mp3", format="mp3")

        # delete file
        os.remove(audioNm)
        os.remove(audioKoNm)

        count = count + 1   
    except: pass

# images to video
path = "C:\\ttsInExcel"
paths = [os.path.join(path , i ) for i in os.listdir(path) if re.search(".png$", i )]

store1 = []

for i in paths :
    store1.append(i)

paths = natsort.natsorted(store1)

pathOut = 'C:\\ttsInExcel\\all_mixed_words.avi'
fps = 0.2

frame_array = []
for idx , path in enumerate(paths) : 
    img = cv2.imread(path)
    height, width, layers = img.shape
    size = (width, height)
    frame_array.append(img)

fourcc = cv2.VideoWriter_fourcc(*'DIVX')

out = cv2.VideoWriter(pathOut,fourcc, fps, size)

for i in range(len(frame_array)):
    # writing to a image array
    out.write(frame_array[i])
out.release()

combined = AudioSegment.empty()
directory = r'C:/ttsInExcel/' 

for file in natsort.natsorted(os.listdir(directory)):
    if file.endswith('.png'): 
        original_path = directory + file
        os.remove(original_path)
    elif file.startswith('audio_mixed') and file.endswith('.mp3'):
        audiofile = AudioSegment.from_mp3(directory + file)
        combined += audiofile
        # save the result
        combined.export("C:\\ttsInExcel\\all_mixed_sounds.mp3", format="mp3")
    if not file.startswith('all_mixed_sounds') and file.endswith('.mp3'):
        original_path = directory + file
        os.remove(original_path)
    
combine_audio("C:\\ttsInExcel\\all_mixed_words.avi","C:\\ttsInExcel\\all_mixed_sounds.mp3","C:\\ttsInExcel\\ttsVideo.mp4")

for file in natsort.natsorted(os.listdir(directory)):
    if file.startswith('all_mixed'):
        original_path = directory + file
        os.remove(original_path)